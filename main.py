# coding=utf-8
# This is a sample Python script.

# Press ⌃R to execute it or replace it with your code.
# Press Double ⇧ to search everywhere for classes, files, tool windows, actions, and settings.
import os
import io
import whatimage
import pyheif
import exifread
from PIL import Image
from PIL.ExifTags import TAGS
from openpyxl import Workbook


def get_exif_data(fname):
    ret = {}
    try:
        img = Image.open(fname)
        if hasattr(img, '_getexif'):
            exifinfo = img._getexif()
            if exifinfo != None:
                for tag, value in exifinfo.items():
                    decoded = TAGS.get(tag, tag)
                    ret[decoded] = value
            else:
                ret = 'no exif'
    except IOError:
        print('IOERROR ' + fname)
    return ret


def get_exif_data_heic(img_path):
    heif_file = pyheif.read(img_path)
    for metadata in heif_file.metadata:
        if metadata['type'] == 'Exif':
            fstream = io.BytesIO(metadata['data'][6:])
            try:
                exifdata = exifread.process_file(fstream, details=False)
                exifdata['DateTime'] = str(exifdata.get('Image DateTime'))
                exifdata['Make'] = str(exifdata.get('Image Make'))
                exifdata['Model'] = str(exifdata.get('Image Model'))
                exifdata['Software'] = str(exifdata.get('Image Software'))
                exifdata['LensModel'] = str(exifdata.get('EXIF LensModel'))
            except UnboundLocalError:
                print("read error: " + img_path)
            return exifdata

    return 'no exif'


def get_location(exif):
    if 'GPSInfo' in exif:
        info = exif.get('GPSInfo')
        if 1 in info and 2 in info:
            lat = info[2]
            lon = info[4]
            lat_d = float(lat[0])
            lat_m = float(lat[1])
            lat_s = float(lat[2])
            lon_d = float(lon[0])
            lon_m = float(lon[1])
            lon_s = float(lon[2])
            lat_decimal = lat_d + (lat_m / 60.0) + (lat_s / 3600.0)
            lon_decimal = lon_d + (lon_m / 60.0) + (lon_s / 3600.0)
            if info[1] is u'S':  # south latitude
                lat_decimal = -lat_decimal
            if info[3] is u'W':  # west longitude
                lon_decimal = -lon_decimal
            return lon_decimal, lat_decimal
    return -999, -999


def get_location_heic(exif):
    if 'GPS GPSLatitude' in exif and 'GPS GPSLongitude' in exif:
        latitude_ref = exif.get('GPS GPSLatitudeRef')
        longitude_ref = exif.get('GPS GPSLongitudeRef')
        latitude_info = exif.get('GPS GPSLatitude').values
        longitude_info = exif.get('GPS GPSLongitude').values
        lat_d = float(latitude_info[0])
        lat_m = float(latitude_info[1])
        lat_s = float(latitude_info[2])
        lon_d = float(longitude_info[0])
        lon_m = float(longitude_info[1])
        lon_s = float(longitude_info[2])
        lat_decimal = lat_d + (lat_m / 60.0) + (lat_s / 3600.0)
        lon_decimal = lon_d + (lon_m / 60.0) + (lon_s / 3600.0)
        if latitude_ref is u'S':  # south latitude
            lat_decimal = -lat_decimal
        if longitude_ref is u'W':  # west longitude
            lon_decimal = -lon_decimal
        return lon_decimal, lat_decimal

    return -999, -999


def obtain_worker():
    table_headers = ['', 'Name', 'Time', 'Device', 'Lens', 'Location']
    wb = Workbook()
    ws = wb.active
    for i in range(1, len(table_headers)):
        ws.cell(row=1, column=i).value = table_headers[i]

    return wb, ws


def is_valid_image(img_path):
    with open(img_path, 'rb') as f:
        data = f.read()
    fmt = whatimage.identify_image(data)
    return fmt


# Press the green button in the gutter to run the script.
if __name__ == '__main__':
    base_dir = './photos'

    valid_format = set(['jpeg', 'heic'])
    for root, dirs, files in os.walk(base_dir):
        workbook, worksheet = obtain_worker()
        index = 2  # exclude header
        succ = True
        for file in files:
            if file.endswith('.DS_Store') or file.endswith('.xlsx'):
                continue

            img_path = os.path.join(root, file)
            fmt = is_valid_image(img_path)
            if not (fmt in valid_format):
                continue

            if 'heic' == fmt:
                exif = get_exif_data_heic(img_path)
            else:
                exif = get_exif_data(img_path)

            if exif == "no exif":
                print('no exif error: ' + img_path)
            else:
                if 'heic' == fmt:
                    lon, lat = get_location_heic(exif)
                else:
                    lon, lat = get_location(exif)

                if lon == -999 and lat == -999:
                    print('location wrong: %s, lat: %d, lon: %d' % (img_path, lat, lon))
                    succ = False
                    continue

                columns = ['']
                columns.append(file)
                columns.append(exif.get('DateTime'))
                columns.append('%s-%s-%s' % (exif.get('Make'), exif.get('Model'), exif.get('Software')))
                columns.append(exif.get('LensModel'))
                columns.append('%s,%s' % (lon, lat))

                for i in range(1, len(columns)):
                    worksheet.cell(row=index, column=i).value = columns[i]

            index = index + 1
        workbook.save(filename='%s.xlsx' % (root))
        if succ == True:
            print('%s done'%(root))

# See PyCharm help at https://www.jetbrains.com/help/pycharm/
